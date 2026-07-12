from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..audit import record_audit_event
from ..config import settings
from ..models import Account, Category, CategoryRule, HoldingSnapshot, ImportBatch, ImportPreset, StagingRow, Transaction
from ..money import parse_decimal_to_cents
from .accounts import infer_account_characterization, infer_last_four, upsert_institution_by_name


CARD_REFERENCE_HEADER = "Posted Date,Reference Number,Payee,Address,Amount"
CHASE_ACTIVITY_HEADER = "Transaction Date,Post Date,Description,Category,Type,Amount,Memo"
CHECKING_HEADER = "Date,Description,Amount,Running Bal."
FIDELITY_HEADER = "Account Number,Account Name,Symbol,Description,Quantity,Last Price,Last Price Change,Current Value"
VENMO_HEADER = ",ID,Datetime,Type,Status,Note,From,To,Amount (total),Amount (tip),Amount (tax),Amount (fee),Tax Rate,Tax Exempt,Funding Source,Destination,Beginning Balance,Ending Balance,Statement Period Venmo Fees,Terminal Location,Year to Date Venmo Fees,Disclaimer"
CITI_ACTIVITY_HEADER = "Status,Date,Description,Debit,Credit"
AMEX_ACTIVITY_HEADER = "Date,Description,Amount"
JPM_POSITIONS_HEADER = "Asset Class,Asset Strategy,Asset Strategy Detail,Description,Ticker,CUSIP,Quantity"


@dataclass
class PreviewResult:
    rows: list[dict]
    warnings: list[str]
    detected_preset: str | None


@dataclass
class AccountImportSuggestion:
    preset_type: str
    suggested_account_id: int | None
    match_confidence: int
    reason: str
    proposed_account: dict
    warnings: list[str]


def decode_text(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("This file is not UTF-8 text. Re-export it as a UTF-8 CSV and try again.") from exc


def detect_preset_from_content(text: str, filename: str = "") -> str | None:
    if CITI_ACTIVITY_HEADER in text:
        normalized_filename = filename.lower()
        return "citi_checking" if normalized_filename.startswith("chk_") or "checking" in normalized_filename else "citi_card_activity"
    for marker, preset in (
        (JPM_POSITIONS_HEADER, "jpm_brokerage_positions"),
        (CARD_REFERENCE_HEADER, "card_reference"),
        (CHASE_ACTIVITY_HEADER, "card_activity"),
        (CHECKING_HEADER, "checking_running_balance"),
        (FIDELITY_HEADER, "brokerage_positions"),
        (VENMO_HEADER, "venmo_activity"),
        (AMEX_ACTIVITY_HEADER, "amex_activity"),
    ):
        if marker in text:
            return preset
    return None


def suggest_account_for_import(db: Session, filename: str, content: bytes) -> AccountImportSuggestion:
    text = decode_text(content)
    preset_type = detect_preset_from_content(text, filename)
    if not preset_type:
        raise ValueError("Could not detect import preset")
    preview = preview_import(content, preset_type)
    proposed = _proposed_account_from_import(filename, preset_type, preview)
    accounts = db.scalars(select(Account).where(Account.status == "active")).all()
    best_account: Account | None = None
    best_score = 0
    best_reason = "No obvious existing account match was found."
    for account in accounts:
        score, reason = _score_account_match(account, proposed, filename)
        if score > best_score:
            best_account = account
            best_score = score
            best_reason = reason
    if best_score < 70:
        exact_candidates = [
            account
            for account in accounts
            if account.account_type == proposed.get("account_type")
            and _account_institution_name(account) == (proposed.get("institution_name") or "").casefold()
        ]
        if len(exact_candidates) == 1:
            best_account = exact_candidates[0]
            best_score = 75
            best_reason = "only active account with matching institution and account type"
    if best_score < 70:
        best_account = None
    return AccountImportSuggestion(
        preset_type=preset_type,
        suggested_account_id=best_account.id if best_account else None,
        match_confidence=best_score if best_account else 0,
        reason=best_reason,
        proposed_account=proposed,
        warnings=preview.warnings,
    )


def parse_csv_preview(content: bytes, preset_type: str) -> PreviewResult:
    text = decode_text(content)
    reader = list(csv.reader(io.StringIO(text)))
    warnings: list[str] = []

    if preset_type == "checking_running_balance":
        header_index = next((i for i, row in enumerate(reader) if row[:4] == ["Date", "Description", "Amount", "Running Bal."]), -1)
        if header_index < 0:
            raise ValueError("Could not find checking ledger header")
        rows = []
        for idx, row in enumerate(reader[header_index + 1 :], start=header_index + 2):
            if not any(cell.strip() for cell in row):
                continue
            if len(row) < 4:
                warnings.append(f"Row {idx} is incomplete")
                continue
            if row[1].startswith("Beginning balance"):
                kind = "balance_marker"
            else:
                kind = "transaction"
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": kind,
                    "transaction_date": row[0],
                    "raw_description": row[1],
                    "amount": row[2],
                    "running_balance": row[3],
                }
            )
        return PreviewResult(rows=rows, warnings=warnings, detected_preset=preset_type)

    header_markers = (
        CARD_REFERENCE_HEADER,
        CHASE_ACTIVITY_HEADER,
        FIDELITY_HEADER,
        VENMO_HEADER,
        CITI_ACTIVITY_HEADER,
        AMEX_ACTIVITY_HEADER,
        JPM_POSITIONS_HEADER,
    )
    header_index = next((i for i, row in enumerate(reader) if any(",".join(row).startswith(marker) for marker in header_markers)), 0)
    data_text = "\n".join(",".join(_csv_escape_cell(cell) for cell in row) for row in reader[header_index:])
    dict_reader = csv.DictReader(io.StringIO(data_text))
    rows = []
    for idx, row in enumerate(dict_reader, start=header_index + 2):
        if not any((value or "").strip() for value in row.values()):
            continue
        row_kind = "transaction"
        if preset_type == "brokerage_positions":
            if row.get("Account Number", "").startswith("Date downloaded") or not row.get("Account Number"):
                continue
            description = row.get("Description") or ""
            if not description and not row.get("Symbol") and not row.get("Current Value"):
                continue
            upper_description = description.upper()
            if upper_description.startswith("BROKERAGELINK") or upper_description.startswith("HELD IN"):
                row_kind = "ignore"
            else:
                row_kind = "position"
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": row_kind,
                    "snapshot_date": None,
                    "account_number": row.get("Account Number"),
                    "symbol": row.get("Symbol"),
                    "description": description,
                    "quantity": row.get("Quantity"),
                    "price": row.get("Last Price"),
                    "market_value": row.get("Current Value"),
                    "asset_class": row.get("Type"),
                    "account_name": row.get("Account Name"),
                }
            )
        elif preset_type == "jpm_brokerage_positions":
            description = (row.get("Description") or "").strip()
            symbol = (row.get("Ticker") or "").strip()
            market_value = (row.get("Value") or "").strip()
            if not (description or symbol) or not market_value or not (row.get("Asset Class") or "").strip():
                continue
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": "position",
                    "snapshot_date": (row.get("As of") or "").strip() or None,
                    "account_number": None,
                    "symbol": symbol,
                    "description": description,
                    "quantity": row.get("Quantity"),
                    "price": row.get("Price"),
                    "market_value": market_value,
                    "asset_class": row.get("Asset Class"),
                    "account_name": "J.P. Morgan Brokerage",
                }
            )
        elif preset_type == "card_activity":
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": row_kind,
                    "transaction_date": row.get("Transaction Date"),
                    "posted_date": row.get("Post Date"),
                    "raw_description": row.get("Description"),
                    "amount": row.get("Amount"),
                    "source_reference": row.get("Memo"),
                    "bank_category": row.get("Category"),
                }
            )
        elif preset_type in {"citi_checking", "citi_card_activity"}:
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": row_kind,
                    "transaction_date": row.get("Date"),
                    "raw_description": row.get("Description"),
                    "amount": _debit_credit_amount(row.get("Debit"), row.get("Credit")),
                }
            )
        elif preset_type == "amex_activity":
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": row_kind,
                    "transaction_date": row.get("Date"),
                    "raw_description": row.get("Description"),
                    "amount": _negate_import_amount(row.get("Amount")),
                }
            )
        elif preset_type == "venmo_activity":
            if not _is_venmo_transaction_row(row):
                continue
            status = (row.get("Status") or "").strip().lower()
            if status and status not in {"complete", "issued"}:
                row_kind = "ignored"
            note = (row.get("Note") or "").strip()
            from_name = (row.get("From") or "").strip()
            to_name = (row.get("To") or "").strip()
            amount = _normalize_import_amount(row.get("Amount (total)"))
            datetime_date = _date_from_iso(row.get("Datetime"))
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": row_kind,
                    "transaction_date": _date_from_venmo_note(note, datetime_date),
                    "posted_date": datetime_date,
                    "raw_description": _venmo_description(note, from_name, to_name, amount),
                    "amount": amount,
                    "source_reference": row.get("ID"),
                    "bank_category": row.get("Type"),
                    "transaction_type": _venmo_transaction_type(row, amount),
                }
            )
        else:
            rows.append(
                {
                    "row_index": idx,
                    "row_kind": row_kind,
                    "transaction_date": row.get("Posted Date"),
                    "raw_description": row.get("Payee"),
                    "amount": row.get("Amount"),
                    "source_reference": row.get("Reference Number"),
                }
            )
    return PreviewResult(rows=rows, warnings=warnings, detected_preset=preset_type)


def _is_venmo_transaction_row(row: dict) -> bool:
    venmo_id = (row.get("ID") or "").strip()
    timestamp = (row.get("Datetime") or "").strip()
    amount = (row.get("Amount (total)") or "").strip()
    return venmo_id.isdigit() and "T" in timestamp and bool(amount)


def _date_from_venmo_note(note: str, fallback_date: str | None) -> str | None:
    fallback = _parse_iso_date(fallback_date)
    year = fallback.year if fallback else date.today().year
    for pattern in (r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", r"\b(\d{1,2})-(\d{1,2})-(\d{2,4})\b"):
        match = re.search(pattern, note)
        if match:
            parsed = _build_date(int(match.group(1)), int(match.group(2)), _expand_year(int(match.group(3))))
            if parsed:
                return _roll_back_future_date(parsed, fallback).isoformat()
    for pattern in (r"\b(\d{1,2})/(\d{1,2})\b", r"\b(\d{1,2})-(\d{1,2})\b"):
        match = re.search(pattern, note)
        if match:
            parsed = _build_date(int(match.group(1)), int(match.group(2)), year)
            if parsed:
                return _roll_back_future_date(parsed, fallback).isoformat()
    return fallback_date

def _roll_back_future_date(candidate: date, reference: date | None) -> date:
    if reference and candidate > reference:
        return date(candidate.year - 1, candidate.month, candidate.day)
    return candidate


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _venmo_transaction_type(row: dict, amount: str | None) -> str:
    amount_cents = parse_decimal_to_cents(amount) or 0
    type_text = (row.get("Type") or "").lower()
    destination = (row.get("Destination") or "").lower()
    funding_source = (row.get("Funding Source") or "").lower()
    if "transfer" in type_text or "standard" in type_text or "bank" in destination or "bank" in funding_source:
        return "transfer"
    if amount_cents > 0:
        return "refund"
    if amount_cents < 0:
        return "expense"
    return "adjustment"

def _venmo_description(note: str, from_name: str, to_name: str, amount: str | None) -> str:
    amount_cents = parse_decimal_to_cents(amount)
    payer, recipient = _venmo_payment_direction(from_name, to_name, amount_cents or 0)
    payment_text = f"{payer} paid {recipient}" if payer and recipient else ""
    parts = [part for part in (note, payment_text) if part]
    return " | ".join(parts) or "Venmo transaction"

def _venmo_payment_direction(from_name: str, to_name: str, amount_cents: int) -> tuple[str | None, str | None]:
    self_name = _venmo_self_name(from_name, to_name)
    if self_name is None:
        # Without a configured owner name (PF_VENMO_SELF_NAME) the money direction is
        # ambiguous for charges, so keep the file's own From -> To order.
        return from_name or None, to_name or None
    other_name = _venmo_other_name(from_name, to_name, self_name)
    if amount_cents < 0:
        return self_name or from_name or None, other_name or to_name or None
    if amount_cents > 0:
        return other_name or to_name or None, self_name or from_name or None
    return from_name or None, to_name or None


def _venmo_self_name(from_name: str, to_name: str) -> str | None:
    configured = (settings.venmo_self_name or "").strip().lower()
    if not configured:
        return None
    for name in (from_name, to_name):
        if name and name.strip().lower() == configured:
            return name
    return None


def _venmo_other_name(from_name: str, to_name: str, self_name: str | None) -> str | None:
    for name in (from_name, to_name):
        if name and name != self_name:
            return name
    return None

def _year_from_date(value: str | None) -> int | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").year
    except ValueError:
        return None


def _expand_year(value: int) -> int:
    if value >= 100:
        return value
    return 2000 + value


def _build_date(month: int, day: int, year: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _csv_escape_cell(value: str) -> str:
    output = io.StringIO()
    csv.writer(output, lineterminator="").writerow([value])
    return output.getvalue()


def _date_from_iso(value: str | None) -> str | None:
    if not value:
        return None
    return value.split("T", 1)[0]


def _normalize_import_amount(value: str | None) -> str | None:
    if value is None:
        return None
    return value.replace("$", "").replace(",", "").replace(" ", "").strip()


def _normalize_transaction_type(description: str, amount_cents: int, account_type: str) -> str:
    text = description.upper()
    if "PAYMENT" in text and account_type == "credit_card":
        return "credit_card_payment"
    if "TRANSFER" in text:
        return "transfer"
    if amount_cents > 0 and account_type in {"checking", "savings"}:
        return "income"
    if amount_cents > 0 and account_type == "credit_card":
        return "refund"
    return "expense"


def _debit_credit_amount(debit: str | None, credit: str | None) -> str | None:
    debit_value = _normalize_import_amount(debit)
    credit_value = _normalize_import_amount(credit)
    if debit_value:
        return _signed_absolute_amount(debit_value, negative=True)
    if credit_value:
        return _signed_absolute_amount(credit_value, negative=False)
    return None


def _negate_import_amount(value: str | None) -> str | None:
    normalized = _normalize_import_amount(value)
    if not normalized:
        return None
    try:
        return format(-Decimal(normalized), "f")
    except InvalidOperation as exc:
        raise ValueError(f"Invalid transaction amount: {value}") from exc


def _signed_absolute_amount(value: str, *, negative: bool) -> str:
    try:
        amount = abs(Decimal(value))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid transaction amount: {value}") from exc
    return format(-amount if negative else amount, "f")


def _proposed_account_from_import(filename: str, preset_type: str, preview: PreviewResult) -> dict:
    account_type = {
        "card_reference": "credit_card",
        "card_activity": "credit_card",
        "checking_running_balance": "checking",
        "brokerage_positions": "brokerage",
        "jpm_brokerage_positions": "brokerage",
        "citi_checking": "checking",
        "citi_card_activity": "credit_card",
        "amex_activity": "credit_card",
        "venmo_activity": "checking",
    }.get(preset_type, "checking")
    institution = _institution_from_filename(filename, preset_type)
    last_four = _last_four_from_import(filename, preview)
    account_name = _account_name_from_preview(preview)
    display_name_parts = [part for part in (institution, account_name, last_four) if part]
    display_name = " ".join(display_name_parts) if display_name_parts else _friendly_name_from_filename(filename)
    return {
        "institution_name": institution,
        "display_name": display_name[:120],
        "account_type": account_type,
        "currency": "USD",
        "last_four": last_four,
    }


def _score_account_match(account: Account, proposed: dict, filename: str) -> tuple[int, str]:
    score = 0
    reasons = []
    proposed_last_four = proposed.get("last_four")
    if proposed_last_four and account.last_four and account.last_four == proposed_last_four:
        score += 70
        reasons.append("last four matched")
    if account.account_type == proposed.get("account_type"):
        score += 20
        reasons.append("account type matched")
    institution_name = account.institution.name if account.institution else None
    proposed_institution = proposed.get("institution_name")
    if institution_name and proposed_institution and institution_name.lower() == proposed_institution.lower():
        score += 25
        reasons.append("institution matched")
    normalized_display = _normalize_account_name(account.display_name)
    normalized_filename = _normalize_account_name(filename)
    normalized_proposed = _normalize_account_name(proposed.get("display_name") or "")
    if normalized_display and (normalized_display in normalized_filename or normalized_display in normalized_proposed):
        score += 15
        reasons.append("name matched")
    return min(score, 100), ", ".join(reasons) or "weak filename/type similarity"



def _account_institution_name(account: Account) -> str:
    if account.institution:
        return account.institution.name.casefold()
    display_name = account.display_name.casefold()
    if "american express" in display_name or "amex" in display_name:
        return "american express"
    if "chase" in display_name or "jpm" in display_name or "j.p. morgan" in display_name:
        return "chase"
    if "citi" in display_name:
        return "citi"
    return ""


def _institution_from_filename(filename: str, preset_type: str) -> str | None:
    text = filename.lower()
    if preset_type == "jpm_brokerage_positions":
        return "Chase"
    if preset_type == "brokerage_positions" or "fidelity" in text or "portfolio_positions" in text or "individual-positions" in text:
        return "Fidelity"
    if preset_type in {"citi_checking", "citi_card_activity"}:
        return "Citi"
    if preset_type == "amex_activity":
        return "American Express"
    if "chase" in text:
        return "Chase"
    if "boa" in text or "bankofamerica" in text or "bank_of_america" in text:
        return "Bank of America"
    if "venmo" in text:
        return "Venmo"
    return None


def _last_four_from_import(filename: str, preview: PreviewResult) -> str | None:
    for row in preview.rows:
        account_number = "".join(char for char in str(row.get("account_number") or "") if char.isdigit())
        if len(account_number) >= 4:
            return account_number[-4:]
    filename_digits = re.findall(r"\d{4,}", filename)
    for token in filename_digits:
        if len(token) == 8 and token.startswith(("19", "20")):
            continue
        if len(token) == 4 and 1900 <= int(token) <= 2099:
            continue
        return token[-4:]
    return None


def _account_name_from_preview(preview: PreviewResult) -> str | None:
    for row in preview.rows:
        value = (row.get("account_name") or "").strip()
        if value:
            return value[:80]
    return None


def _friendly_name_from_filename(filename: str) -> str:
    stem = filename.rsplit(".", 1)[0]
    cleaned = re.sub(r"[_\-]+", " ", stem).strip()
    return cleaned.title() or "Imported Account"


def _source_hash(account_id: int, date_value: str, amount: str, description: str, source_reference: str | None, ordinal: int) -> str:
    payload = "|".join([str(account_id), date_value or "", amount or "", description or "", source_reference or "", str(ordinal)])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()

def _history_source_hash(account_id: int, date_value: str, amount: str, description: str, category: str, ordinal: int) -> str:
    payload = "|".join(["categorized_history", str(account_id), date_value, amount, description, category, str(ordinal)])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _category_key_from_label(label: str) -> str:
    key = "".join(char.lower() if char.isalnum() else "_" for char in label.strip())
    key = "_".join(part for part in key.split("_") if part)
    return key[:60] or "category"


def _unique_category_key(db: Session, label: str) -> str:
    base = _category_key_from_label(label)
    key = base
    suffix = 2
    while db.scalar(select(Category).where(Category.key == key)):
        key = f"{base[:55]}_{suffix}"
        suffix += 1
    return key


def _find_or_create_category(db: Session, label: str | None) -> tuple[Category | None, bool]:
    cleaned = (label or "").strip()
    if not cleaned:
        return None, False
    existing = db.scalar(select(Category).where(Category.label == cleaned))
    if existing:
        return existing, False
    category = Category(key=_unique_category_key(db, cleaned), label=cleaned)
    db.add(category)
    db.flush()
    return category, True


def _find_or_create_history_account(db: Session, account_name: str) -> tuple[Account, bool]:
    cleaned = account_name.strip()
    if not cleaned:
        raise ValueError("Categorized history rows must include an Account value")
    existing = db.scalar(select(Account).where(Account.display_name == cleaned))
    if not existing:
        existing = db.scalar(select(Account).where(Account.display_name.ilike(cleaned)))
    if existing:
        if existing.status != "active":
            existing.status = "active"
        return existing, False
    characterization = infer_account_characterization(cleaned)
    institution = upsert_institution_by_name(db, characterization.institution_name)
    account = Account(
        institution_id=institution.id if institution else None,
        display_name=characterization.display_name or cleaned,
        account_type=characterization.account_type,
        currency="USD",
        status="active",
        last_four=infer_last_four(characterization.display_name or cleaned),
    )
    db.add(account)
    db.flush()
    return account, True


def _spreadsheet_cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _read_history_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else "csv"
    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [[_spreadsheet_cell_to_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    else:
        text = decode_text(content)
        rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise ValueError("The categorized history file is empty")

    header_index = None
    normalized_headers: list[str] = []
    for index, row in enumerate(rows[:20]):
        normalized = [str(value).strip().lower() for value in row]
        if {"account", "posted date", "payee", "amount"}.issubset(set(normalized)):
            header_index = index
            normalized_headers = normalized
            break
    if header_index is None:
        raise ValueError('Could not find columns named "Account", "Posted Date", "Payee", and "Amount"')

    def column(name: str) -> int:
        return normalized_headers.index(name)

    account_index = column("account")
    date_index = column("posted date")
    payee_index = column("payee")
    amount_index = column("amount")
    category_index = normalized_headers.index("expense category") if "expense category" in normalized_headers else None

    parsed: list[dict[str, str]] = []
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = [str(value).strip() for value in row]
        if not any(values):
            continue
        get = lambda index: values[index] if index < len(values) else ""
        parsed.append(
            {
                "row_index": str(row_index),
                "account": get(account_index),
                "posted_date": get(date_index),
                "payee": get(payee_index),
                "amount": get(amount_index),
                "category": get(category_index) if category_index is not None else "",
            }
        )
    return parsed





def _history_row_errors(row: dict[str, str]) -> list[str]:
    errors: list[str] = []
    if not (row.get("account") or "").strip():
        errors.append("Account")
    posted_date_text = (row.get("posted_date") or "").strip()
    if not posted_date_text:
        errors.append("Posted Date")
    else:
        try:
            _parse_import_date(posted_date_text)
        except ValueError:
            errors.append("Posted Date")
    if not (row.get("payee") or "").strip():
        errors.append("Payee")
    amount_text = (row.get("amount") or "").strip()
    if not amount_text:
        errors.append("Amount")
    else:
        try:
            parse_decimal_to_cents(amount_text)
        except ValueError:
            errors.append("Amount")
    return errors


def review_categorized_history(filename: str, content: bytes) -> dict:
    rows = _read_history_rows(filename, content)
    reviewed_rows = [{**row, "errors": _history_row_errors(row)} for row in rows]
    return {"rows": reviewed_rows, "needs_review": any(row["errors"] for row in reviewed_rows)}


def _commit_categorized_history_rows(db: Session, filename: str, rows: list[dict[str, str]], actor: str = "local-user") -> dict:
    if not rows:
        raise ValueError("The categorized history file did not contain transaction rows")
    invalid_rows = [{**row, "errors": _history_row_errors(row)} for row in rows if _history_row_errors(row)]
    if invalid_rows:
        raise ValueError("Some categorized history rows still need Account, Posted Date, Payee, or Amount before import")

    file_hash = hashlib.sha256(json.dumps(rows, sort_keys=True).encode("utf-8")).hexdigest()
    batches_by_account_id: dict[int, ImportBatch] = {}
    imported_by_account_id: Counter[int] = Counter()
    skipped_by_account_id: Counter[int] = Counter()
    description_counter: Counter[tuple[int, str, str, str, str]] = Counter()
    inserted = 0
    skipped = 0
    accounts_created: set[int] = set()
    categories_created: set[int] = set()
    warnings: list[str] = []

    for row in rows:
        account, account_created = _find_or_create_history_account(db, row["account"])
        if account_created:
            accounts_created.add(account.id)
        category, category_created = _find_or_create_category(db, row.get("category"))
        if category and category_created:
            categories_created.add(category.id)

        if account.id not in batches_by_account_id:
            batch = ImportBatch(account_id=account.id, preset_id=None, filename=filename, file_hash=file_hash, status="committed")
            db.add(batch)
            db.flush()
            batches_by_account_id[account.id] = batch
        batch = batches_by_account_id[account.id]

        posted_date_text = (row.get("posted_date") or "").strip()
        amount_text = (row.get("amount") or "").strip()
        description = (row.get("payee") or "").strip()
        transaction_date = _parse_import_date(posted_date_text)
        amount_cents = parse_decimal_to_cents(amount_text) or 0

        key = (account.id, transaction_date.isoformat(), amount_text, description, row.get("category") or "")
        description_counter[key] += 1
        ordinal = description_counter[key]
        source_hash = _history_source_hash(account.id, transaction_date.isoformat(), amount_text, description, row.get("category") or "", ordinal)
        existing = db.scalar(select(Transaction).where(Transaction.account_id == account.id, Transaction.source_hash == source_hash))
        if existing:
            skipped += 1
            skipped_by_account_id[account.id] += 1
            continue

        db.add(
            StagingRow(
                import_batch_id=batch.id,
                account_id=account.id,
                row_index=int(row["row_index"]),
                row_kind="transaction",
                raw_json=json.dumps(row, default=str),
                normalized_json=json.dumps(row, default=str),
            )
        )
        db.add(
            Transaction(
                account_id=account.id,
                import_batch_id=batch.id,
                transaction_date=transaction_date,
                posted_date=transaction_date,
                amount_cents=amount_cents,
                raw_description=description,
                normalized_payee=description[:255],
                transaction_type=_history_transaction_type(row.get("category"), amount_cents, account.account_type),
                category_id=category.id if category else None,
                review_status="confirmed",
                source_hash=source_hash,
                source_reference=f"categorized-history-row-{row['row_index']}",
                source_ordinal=ordinal,
            )
        )
        inserted += 1
        imported_by_account_id[account.id] += 1

    for account_id, batch in batches_by_account_id.items():
        batch.imported_rows = imported_by_account_id[account_id]
        batch.skipped_duplicates = skipped_by_account_id[account_id]
        batch.warnings_json = json.dumps(warnings)

    record_audit_event(
        db,
        "categorized_history_import",
        actor,
        "import_batch",
        filename,
        {"filename": filename, "inserted": inserted, "skipped": skipped, "accounts_created": len(accounts_created), "categories_created": len(categories_created)},
    )
    return {
        "inserted": inserted,
        "skipped": skipped,
        "accounts_created": len(accounts_created),
        "categories_created": len(categories_created),
        "warnings": warnings,
    }

def _history_transaction_type(category_label: str | None, amount_cents: int, account_type: str) -> str:
    label = (category_label or "").strip().lower()
    if "income" in label:
        return "income"
    if "transfer" in label or "payment" in label:
        return "transfer"
    if "refund" in label:
        return "refund"
    if amount_cents > 0 and account_type == "credit_card":
        return "expense"
    if amount_cents > 0 and account_type in {"checking", "savings", "cash"}:
        return "income"
    return "expense"


def commit_categorized_history(db: Session, filename: str, content: bytes, actor: str = "local-user") -> dict:
    rows = _read_history_rows(filename, content)
    return _commit_categorized_history_rows(db, filename, rows, actor)


def commit_reviewed_categorized_history(db: Session, filename: str, rows: list[dict[str, str]], actor: str = "local-user") -> dict:
    cleaned_rows = [
        {
            "row_index": str(row.get("row_index") or index + 1),
            "account": str(row.get("account") or "").strip(),
            "posted_date": str(row.get("posted_date") or "").strip(),
            "payee": str(row.get("payee") or "").strip(),
            "amount": str(row.get("amount") or "").strip(),
            "category": str(row.get("category") or "").strip(),
        }
        for index, row in enumerate(rows)
    ]
    return _commit_categorized_history_rows(db, filename, cleaned_rows, actor)


def preview_import(content: bytes, preset_type: str) -> PreviewResult:
    return parse_csv_preview(content, preset_type)


def commit_import(db: Session, account, preset: ImportPreset | None, filename: str, content: bytes, actor: str = "local-user", snapshot_date: date | None = None) -> dict:
    detected = preset.preset_type if preset else detect_preset_from_content(decode_text(content), filename)
    if not detected:
        raise ValueError("Unable to detect preset type")
    preview = preview_import(content, detected)
    file_hash = hashlib.sha256(content).hexdigest()
    batch = ImportBatch(account_id=account.id, preset_id=preset.id if preset else None, filename=filename, file_hash=file_hash, status="committed")
    db.add(batch)
    db.flush()

    inserted = 0
    skipped = 0
    warnings = list(preview.warnings)
    description_counter: Counter[tuple[str, str, str | None]] = Counter()
    rules = db.scalars(select(CategoryRule).order_by(CategoryRule.priority.asc(), CategoryRule.id.asc())).all()

    resolved_snapshot_date: date | None = None
    cleared_snapshot_scopes: set[tuple[int, str]] = set()
    if _is_brokerage_preset(detected):
        resolved_snapshot_date = snapshot_date or _snapshot_date_from_preview(preview) or _extract_snapshot_date(filename)
        if resolved_snapshot_date is None:
            resolved_snapshot_date = date.today()
            warnings.append(
                f"Could not find a date in the filename \"{filename}\"; recorded these positions as of today ({resolved_snapshot_date.isoformat()})."
            )

    for row in preview.rows:
        target_account = account
        if _is_brokerage_preset(detected):
            target_account, routing_warning = _resolve_brokerage_account(db, account, row)
            if routing_warning and routing_warning not in warnings:
                warnings.append(routing_warning)
        db.add(
            StagingRow(
                import_batch_id=batch.id,
                account_id=target_account.id,
                row_index=row["row_index"],
                row_kind=row["row_kind"],
                raw_json=json.dumps(row, default=str),
                normalized_json=json.dumps(row, default=str),
            )
        )
        if _is_brokerage_preset(detected):
            if row["row_kind"] == "ignore":
                continue
            market_value_cents = parse_decimal_to_cents(row.get("market_value"))
            if market_value_cents is not None:
                scope = (target_account.id, resolved_snapshot_date.isoformat())
                if scope not in cleared_snapshot_scopes:
                    # Re-importing a positions file replaces that account/date snapshot
                    # instead of double-counting it.
                    stale = db.scalars(
                        select(HoldingSnapshot).where(
                            HoldingSnapshot.account_id == target_account.id,
                            HoldingSnapshot.snapshot_date == resolved_snapshot_date,
                        )
                    ).all()
                    for stale_row in stale:
                        db.delete(stale_row)
                    if stale:
                        db.flush()
                    cleared_snapshot_scopes.add(scope)
                db.add(
                    HoldingSnapshot(
                        account_id=target_account.id,
                        snapshot_date=resolved_snapshot_date,
                        symbol=row.get("symbol"),
                        description=row.get("description"),
                        quantity_basis_points=_parse_decimal_to_basis_points(row.get("quantity")),
                        price_cents=parse_decimal_to_cents(row.get("price")),
                        market_value_cents=market_value_cents,
                        asset_class=row.get("asset_class"),
                    )
                )
                inserted += 1
            continue

        if row["row_kind"] != "transaction":
            continue

        key = (row.get("transaction_date") or "", row.get("amount") or "", row.get("raw_description") or "")
        description_counter[key] += 1
        ordinal = description_counter[key]
        source_hash = _source_hash(
            account.id,
            row.get("transaction_date") or "",
            row.get("amount") or "",
            row.get("raw_description") or "",
            row.get("source_reference"),
            ordinal,
        )
        existing = db.scalar(select(Transaction).where(Transaction.account_id == account.id, Transaction.source_hash == source_hash))
        if existing:
            skipped += 1
            continue

        amount_cents = parse_decimal_to_cents(row.get("amount")) or 0
        review_status = "needs_review"
        category_id = None
        rule_transaction_type: str | None = None
        normalized = (row.get("raw_description") or "").strip()
        haystack = normalized.upper()
        for rule in rules:
            if rule.field_name == "raw_description" and rule.match_text.upper() in haystack:
                category_id = rule.category_id
                rule_transaction_type = rule.suggested_transaction_type
                review_status = "suggested"
                break

        transaction = Transaction(
            account_id=account.id,
            import_batch_id=batch.id,
            transaction_date=_parse_import_date(row.get("transaction_date")),
            posted_date=_parse_import_date(row.get("posted_date")) if row.get("posted_date") else None,
            amount_cents=amount_cents,
            raw_description=normalized,
            normalized_payee=normalized[:255],
            transaction_type=rule_transaction_type or row.get("transaction_type") or _normalize_transaction_type(normalized, amount_cents, account.account_type),
            category_id=category_id,
            review_status=review_status,
            source_hash=source_hash,
            source_reference=row.get("source_reference"),
            source_ordinal=ordinal,
            running_balance_cents=parse_decimal_to_cents(row.get("running_balance")),
        )
        duplicate_of_id = _find_possible_duplicate_id(db, account.id, transaction)
        if duplicate_of_id is not None:
            transaction.review_status = "possible_duplicate"
            transaction.duplicate_of_transaction_id = duplicate_of_id
        db.add(transaction)
        inserted += 1

    batch.imported_rows = inserted
    batch.skipped_duplicates = skipped
    batch.warnings_json = json.dumps(warnings)
    record_audit_event(db, "import_commit", actor, "import_batch", str(batch.id), {"filename": filename, "inserted": inserted, "skipped": skipped})
    return {"batch_id": batch.id, "inserted": inserted, "skipped": skipped, "warnings": warnings}


def _resolve_brokerage_account(db: Session, selected_account: Account, row: dict) -> tuple[Account, str | None]:
    account_name = (row.get("account_name") or "").strip()
    account_number = "".join(char for char in (row.get("account_number") or "") if char.isdigit())
    query = select(Account).where(Account.status == "active")
    if selected_account.institution_id:
        query = query.where(Account.institution_id == selected_account.institution_id)
    else:
        query = query.where(Account.id == selected_account.id)
    candidates = db.scalars(query).all()

    for candidate in candidates:
        if candidate.last_four and account_number.endswith(candidate.last_four):
            return candidate, None

    normalized_row_name = _normalize_account_name(account_name)
    for candidate in candidates:
        normalized_display = _normalize_account_name(candidate.display_name)
        if normalized_display and normalized_row_name and (normalized_display in normalized_row_name or normalized_row_name in normalized_display):
            return candidate, None

    label = account_name or row.get("account_number") or "unknown account"
    return selected_account, f'Could not match brokerage account "{label}"; assigned those holdings to the selected account "{selected_account.display_name}".'


def _normalize_account_name(value: str) -> str:
    return "".join(char.lower() for char in value if char.isalnum())

def _is_brokerage_preset(preset_type: str) -> bool:
    return preset_type in {"brokerage_positions", "jpm_brokerage_positions"}


def _snapshot_date_from_preview(preview: PreviewResult) -> date | None:
    for row in preview.rows:
        value = row.get("snapshot_date")
        if value:
            return _parse_import_date(value)
    return None



def _parse_decimal_to_basis_points(value: str | int | float | None) -> int | None:
    if value is None or value == "":
        return None
    try:
        normalized = str(value).replace(",", "").strip()
        return int((Decimal(normalized) * Decimal("10000")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid quantity value: {value}") from exc


def _parse_import_date(value: str | None) -> date:
    if not value:
        raise ValueError("Import row is missing a transaction date")
    for date_format in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported import date: {value}")




def _find_possible_duplicate_id(db: Session, account_id: int, candidate: Transaction) -> int | None:
    existing = db.scalar(
        select(Transaction.id).where(
            Transaction.account_id == account_id,
            Transaction.transaction_date == candidate.transaction_date,
            Transaction.amount_cents == candidate.amount_cents,
            Transaction.raw_description != candidate.raw_description,
        ).limit(1)
    )
    return existing


MONTH_ABBREVIATIONS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _extract_snapshot_date(filename: str) -> date | None:
    """Parse a snapshot date out of a positions filename, or return None when absent.

    Handles the common export shapes:
    - Portfolio_Positions_Jul-04-2026.csv (Fidelity style)
    - positions-2026-07-04.csv / positions_20260704.csv
    - positions 07-04-2026.csv
    """
    text = filename.lower()
    match = re.search(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[-_ ](\d{1,2})[-_ ](\d{4})", text)
    if match:
        parsed = _build_date(MONTH_ABBREVIATIONS[match.group(1)], int(match.group(2)), int(match.group(3)))
        if parsed:
            return parsed
    match = re.search(r"(\d{4})[-_ .](\d{1,2})[-_ .](\d{1,2})", text)
    if match:
        parsed = _build_date(int(match.group(2)), int(match.group(3)), int(match.group(1)))
        if parsed:
            return parsed
    match = re.search(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)", text)
    if match:
        parsed = _build_date(int(match.group(2)), int(match.group(3)), int(match.group(1)))
        if parsed:
            return parsed
    match = re.search(r"(?<!\d)(\d{1,2})[-_ .](\d{1,2})[-_ .](\d{4})(?!\d)", text)
    if match:
        parsed = _build_date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        if parsed:
            return parsed
    return None
